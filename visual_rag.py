import os
import sys
import glob
import io
import numpy as np
import torch
import open_clip
import faiss

from torchvision import datasets, transforms
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

# ── Config ────────────────────────────────────────────────────────────────────
NUM_IMAGES   = 2500
BATCH_SIZE   = 64
TOP_K        = 10
CLIP_MODEL   = "ViT-B-32"
PRETRAINED   = "openai"
DATA_DIR     = "./cifar_data"
QUERY_FOLDER = "./querries"      # folder with your query images
RESULTS_DIR  = "./results"

IMG_SIZE     = 80                # resize all images to 80×80 px
COL_WIDTH    = 12                # Excel column width units
ROW_HEIGHT   = 65                # Excel row height in points

CLASSES = [
    "airplane", "automobile", "bird", "cat", "deer",
    "dog", "frog", "horse", "ship", "truck"
]

# ── ML helpers ────────────────────────────────────────────────────────────────

def load_cifar10():
    print(f"\n[1/5] Loading CIFAR-10 ({NUM_IMAGES} images)...")
    dataset = datasets.CIFAR10(
        root=DATA_DIR, train=True, download=True,
        transform=transforms.ToTensor()
    )
    images_tensor, labels = [], []
    for idx in range(NUM_IMAGES):
        img, label = dataset[idx]
        images_tensor.append(img)
        labels.append(label)
    print(f"    Loaded {len(images_tensor)} images.")
    return images_tensor, labels


def load_clip_model():
    print(f"\n[2/5] Loading CLIP model...")
    device = "cuda" if torch.cuda.is_available() else "cpu"
    print(f"    Using device: {device}")
    model, _, preprocess = open_clip.create_model_and_transforms(
        CLIP_MODEL, pretrained=PRETRAINED
    )
    model = model.to(device).eval()
    return model, preprocess, device


def tensor_to_pil(img_tensor):
    np_img = (img_tensor.permute(1, 2, 0).numpy() * 255).astype(np.uint8)
    return Image.fromarray(np_img)


def generate_embeddings(images_tensor, model, preprocess, device):
    print(f"\n[3/5] Generating embeddings...")
    all_embeddings = []
    with torch.no_grad():
        for start in range(0, len(images_tensor), BATCH_SIZE):
            batch = images_tensor[start:start + BATCH_SIZE]
            clip_input = torch.stack(
                [preprocess(tensor_to_pil(t)) for t in batch]
            ).to(device)
            emb = model.encode_image(clip_input)
            emb = emb / emb.norm(dim=-1, keepdim=True)
            all_embeddings.append(emb.cpu().numpy())
            done = min(start + BATCH_SIZE, len(images_tensor))
            print(f"    {done}/{len(images_tensor)}", end="\r")
    print()
    embeddings_np = np.vstack(all_embeddings).astype(np.float32)
    print(f"    Shape: {embeddings_np.shape}")
    return embeddings_np


def build_faiss_index(embeddings):
    print(f"\n[4/5] Building FAISS index...")
    dim   = embeddings.shape[1]
    index = faiss.IndexFlatIP(dim)
    index.add(embeddings)
    print(f"    Vectors stored: {index.ntotal}")
    return index


def query_image(image_path, model, preprocess, device, index):
    pil_img    = Image.open(image_path).convert("RGB")
    clip_input = preprocess(pil_img).unsqueeze(0).to(device)
    with torch.no_grad():
        emb = model.encode_image(clip_input)
        emb = emb / emb.norm(dim=-1, keepdim=True)
        emb = emb.cpu().numpy().astype(np.float32)
    distances, indices = index.search(emb, TOP_K)
    return distances[0], indices[0], pil_img


# ── Image → openpyxl helper ───────────────────────────────────────────────────

def pil_to_xl_image(pil_img, size=IMG_SIZE):
    resized = pil_img.resize((size, size), Image.LANCZOS)
    buf = io.BytesIO()
    resized.save(buf, format="PNG")
    buf.seek(0)
    xl_img = XLImage(buf)
    xl_img.width  = size
    xl_img.height = size
    return xl_img


# ── Excel style helpers ───────────────────────────────────────────────────────

HEADER_FILL  = PatternFill("solid", start_color="1F4E79")
QUERY_FILL   = PatternFill("solid", start_color="FFF2CC")
HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
LABEL_FONT   = Font(name="Arial", bold=True, size=9)
SCORE_FONT   = Font(name="Arial", size=8, italic=True)
CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN         = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin")
)

def set_cell(ws, row, col, value=None, font=None, fill=None):
    cell = ws.cell(row=row, column=col)
    if value is not None:
        cell.value = value
    cell.font      = font or Font(name="Arial", size=9)
    if fill:
        cell.fill  = fill
    cell.border    = THIN
    cell.alignment = CENTER
    return cell


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    if not os.path.isdir(QUERY_FOLDER):
        print(f"ERROR: Query folder not found — '{QUERY_FOLDER}'")
        sys.exit(1)

    exts = ("*.jpg", "*.jpeg", "*.png", "*.bmp", "*.webp")
    query_paths = []
    for ext in exts:
        query_paths.extend(glob.glob(os.path.join(QUERY_FOLDER, ext)))
        query_paths.extend(glob.glob(os.path.join(QUERY_FOLDER, ext.upper())))
    query_paths = sorted(set(query_paths))

    if not query_paths:
        print(f"ERROR: No images found in '{QUERY_FOLDER}'")
        sys.exit(1)

    print("=" * 55)
    print("  Batch Visual RAG  —  CIFAR-10 + CLIP + FAISS")
    print("=" * 55)
    print(f"  Query images found : {len(query_paths)}")

    os.makedirs(RESULTS_DIR, exist_ok=True)

    images_tensor, labels     = load_cifar10()
    model, preprocess, device = load_clip_model()
    embeddings                = generate_embeddings(images_tensor, model, preprocess, device)
    index                     = build_faiss_index(embeddings)

    wb = Workbook()

    # ── Sheet 1: Overview (text only) ─────────────────────────────────────────
    overview = wb.active
    overview.title = "Overview"
    ov_headers = ["Query #", "Filename", "Top-1 Class", "Top-1 Score",
                  "Top-3 Classes", "Top-3 Scores", "Top-10 Classes", "Top-10 Scores"]
    overview.append(ov_headers)
    for col in range(1, len(ov_headers) + 1):
        set_cell(overview, 1, col, font=HEADER_FONT, fill=HEADER_FILL)
    for i, w in enumerate([8, 28, 14, 12, 38, 32, 70, 65], 1):
        overview.column_dimensions[get_column_letter(i)].width = w
    overview.freeze_panes = "A2"
    overview.auto_filter.ref = f"A1:{get_column_letter(len(ov_headers))}1"

    # ── Sheet 2: Visual Results (images) ──────────────────────────────────────
    visual = wb.create_sheet("Visual Results")

    # Header row
    set_cell(visual, 1, 1, "Query Image", font=HEADER_FONT, fill=HEADER_FILL)
    set_cell(visual, 1, 2, "Filename",    font=HEADER_FONT, fill=HEADER_FILL)
    for rank in range(1, TOP_K + 1):
        set_cell(visual, 1, rank + 2, f"#{rank}", font=HEADER_FONT, fill=HEADER_FILL)

    visual.column_dimensions["A"].width = COL_WIDTH
    visual.column_dimensions["B"].width = 22
    for col in range(3, TOP_K + 4):
        visual.column_dimensions[get_column_letter(col)].width = COL_WIDTH

    print(f"\n[5/5] Running {len(query_paths)} queries...\n")

    for q_idx, img_path in enumerate(query_paths, 1):
        fname = os.path.basename(img_path)
        print(f"  [{q_idx:>3}/{len(query_paths)}] {fname}", end=" ... ")

        try:
            distances, indices, query_pil = query_image(
                img_path, model, preprocess, device, index
            )
        except Exception as e:
            print(f"FAILED ({e})")
            continue

        result_labels  = [labels[i]  for i in indices]
        result_classes = [CLASSES[l] for l in result_labels]
        result_pils    = [tensor_to_pil(images_tensor[i]) for i in indices]

        # Each query = 3 rows: [image row] [class label row] [score row]
        img_row   = (q_idx - 1) * 3 + 2
        label_row = img_row + 1
        score_row = img_row + 2

        visual.row_dimensions[img_row].height   = ROW_HEIGHT
        visual.row_dimensions[label_row].height = 16
        visual.row_dimensions[score_row].height = 14

        # Col A — query image
        visual.add_image(pil_to_xl_image(query_pil), f"A{img_row}")
        set_cell(visual, img_row,   1, fill=QUERY_FILL)
        set_cell(visual, label_row, 1, "INPUT", font=LABEL_FONT, fill=QUERY_FILL)
        set_cell(visual, score_row, 1, fill=QUERY_FILL)

        # Col B — filename (merged across 3 rows)
        visual.merge_cells(f"B{img_row}:B{score_row}")
        set_cell(visual, img_row, 2, fname,
                 font=Font(name="Arial", bold=True, size=9), fill=QUERY_FILL)

        # Cols C+ — result images, labels, scores
        for rank, (pil_img, cls, dist) in enumerate(
            zip(result_pils, result_classes, distances), 1
        ):
            col        = rank + 2
            col_letter = get_column_letter(col)
            alt_fill   = PatternFill("solid", start_color="EBF5FB") if rank % 2 == 0 else None

            visual.add_image(pil_to_xl_image(pil_img), f"{col_letter}{img_row}")
            set_cell(visual, img_row,   col, fill=alt_fill)
            set_cell(visual, label_row, col, cls,        font=LABEL_FONT,  fill=alt_fill)
            set_cell(visual, score_row, col, f"{dist:.4f}", font=SCORE_FONT, fill=alt_fill)

        # Overview row
        ov_row  = q_idx + 1
        alt_ov  = PatternFill("solid", start_color="D6E4F0") if q_idx % 2 == 0 else None
        ov_data = [
            q_idx, fname,
            result_classes[0], round(float(distances[0]), 4),
            ", ".join(result_classes[:3]),
            ", ".join(f"{d:.4f}" for d in distances[:3]),
            ", ".join(result_classes),
            ", ".join(f"{d:.4f}" for d in distances),
        ]
        overview.append(ov_data)
        for col in range(1, len(ov_data) + 1):
            set_cell(overview, ov_row, col, fill=alt_ov)

        print(f"Top-1 → {result_classes[0]} ({distances[0]:.4f})")

    excel_path = os.path.join(RESULTS_DIR, "visual_rag_results.xlsx")
    wb.save(excel_path)
    print(f"\n  Excel saved → '{excel_path}'")
    print(f"  Sheets: 'Overview' (text summary) + 'Visual Results' (embedded images)")
    print("\n Done!\n")


if __name__ == "__main__":
    main()





# import os
# import sys
# import glob
# import numpy as np
# import torch
# import open_clip
# import faiss
# import shutil

# from torchvision import datasets, transforms
# from PIL import Image
# from openpyxl import Workbook
# from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# from openpyxl.utils import get_column_letter

# # ── Config ──────────────────────────────────────────────────────────────────
# NUM_IMAGES  = 2000
# BATCH_SIZE  = 64
# TOP_K       = 10
# CLIP_MODEL  = "ViT-B-32"
# PRETRAINED  = "openai"
# DATA_DIR    = "./cifar_data"

# # ── Put all 50 query images in this folder ──────────────────────────────────
# QUERY_FOLDER = "./querries"   # <-- place your 50 images here

# RESULTS_DIR  = "./results"        # output folder

# CLASSES = [
#     "airplane", "automobile", "bird", "cat", "deer",
#     "dog", "frog", "horse", "ship", "truck"
# ]

# # ── Helpers ──────────────────────────────────────────────────────────────────

# def load_cifar10():
#     print(f"\n[1/5] Loading CIFAR-10 ({NUM_IMAGES} images)...")
#     dataset = datasets.CIFAR10(
#         root=DATA_DIR, train=True, download=True,
#         transform=transforms.ToTensor()
#     )
#     images_tensor, labels = [], []
#     for idx in range(NUM_IMAGES):
#         img, label = dataset[idx]
#         images_tensor.append(img)
#         labels.append(label)
#     print(f"    Loaded {len(images_tensor)} images.")
#     return images_tensor, labels


# def load_clip_model():
#     print(f"\n[2/5] Loading CLIP model...")
#     device = "cuda" if torch.cuda.is_available() else "cpu"
#     print(f"    Using device: {device}")
#     model, _, preprocess = open_clip.create_model_and_transforms(
#         CLIP_MODEL, pretrained=PRETRAINED
#     )
#     model = model.to(device).eval()
#     return model, preprocess, device


# def tensor_to_pil(img_tensor):
#     np_img = (img_tensor.permute(1, 2, 0).numpy() * 255).astype(np.uint8)
#     return Image.fromarray(np_img)


# def generate_embeddings(images_tensor, model, preprocess, device):
#     print(f"\n[3/5] Generating embeddings...")
#     all_embeddings = []
#     with torch.no_grad():
#         for start in range(0, len(images_tensor), BATCH_SIZE):
#             batch = images_tensor[start:start + BATCH_SIZE]
#             clip_input = torch.stack(
#                 [preprocess(tensor_to_pil(t)) for t in batch]
#             ).to(device)
#             emb = model.encode_image(clip_input)
#             emb = emb / emb.norm(dim=-1, keepdim=True)
#             all_embeddings.append(emb.cpu().numpy())
#             done = min(start + BATCH_SIZE, len(images_tensor))
#             print(f"    {done}/{len(images_tensor)}", end="\r")
#     print()
#     embeddings_np = np.vstack(all_embeddings).astype(np.float32)
#     print(f"    Shape: {embeddings_np.shape}")
#     return embeddings_np


# def build_faiss_index(embeddings):
#     print(f"\n[4/5] Building FAISS index...")
#     dim   = embeddings.shape[1]
#     index = faiss.IndexFlatIP(dim)
#     index.add(embeddings)
#     print(f"    Vectors stored: {index.ntotal}")
#     return index


# def query_image(image_path, model, preprocess, device, index):
#     pil_img    = Image.open(image_path).convert("RGB")
#     clip_input = preprocess(pil_img).unsqueeze(0).to(device)
#     with torch.no_grad():
#         emb = model.encode_image(clip_input)
#         emb = emb / emb.norm(dim=-1, keepdim=True)
#         emb = emb.cpu().numpy().astype(np.float32)
#     distances, indices = index.search(emb, TOP_K)
#     return distances[0], indices[0]


# # ── Excel helpers ─────────────────────────────────────────────────────────────

# HEADER_FILL   = PatternFill("solid", start_color="1F4E79")
# ALT_ROW_FILL  = PatternFill("solid", start_color="D6E4F0")
# HEADER_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
# BODY_FONT     = Font(name="Arial", size=10)
# CENTER        = Alignment(horizontal="center", vertical="center")
# THIN_BORDER   = Border(
#     left=Side(style="thin"), right=Side(style="thin"),
#     top=Side(style="thin"),  bottom=Side(style="thin")
# )

# def style_header(ws, row, cols):
#     for col in range(1, cols + 1):
#         cell = ws.cell(row=row, column=col)
#         cell.fill   = HEADER_FILL
#         cell.font   = HEADER_FONT
#         cell.alignment = CENTER
#         cell.border = THIN_BORDER

# def style_body_row(ws, row, cols, alternate=False):
#     for col in range(1, cols + 1):
#         cell = ws.cell(row=row, column=col)
#         if alternate:
#             cell.fill = ALT_ROW_FILL
#         cell.font      = BODY_FONT
#         cell.alignment = CENTER
#         cell.border    = THIN_BORDER


# # ── Main ──────────────────────────────────────────────────────────────────────

# def main():
#     # ── Validate query folder ────────────────────────────────────────────────
#     if not os.path.isdir(QUERY_FOLDER):
#         print(f"ERROR: Query folder not found — '{QUERY_FOLDER}'")
#         print("Create the folder and place your query images inside it.")
#         sys.exit(1)

#     exts = ("*.jpg", "*.jpeg", "*.png", "*.bmp", "*.webp")
#     query_paths = []
#     for ext in exts:
#         query_paths.extend(glob.glob(os.path.join(QUERY_FOLDER, ext)))
#         query_paths.extend(glob.glob(os.path.join(QUERY_FOLDER, ext.upper())))
#     query_paths = sorted(set(query_paths))

#     if not query_paths:
#         print(f"ERROR: No images found in '{QUERY_FOLDER}'")
#         sys.exit(1)

#     print("=" * 55)
#     print("  Batch Visual RAG  —  CIFAR-10 + CLIP + FAISS")
#     print("=" * 55)
#     print(f"  Query images found : {len(query_paths)}")

#     # ── Setup ────────────────────────────────────────────────────────────────
#     os.makedirs(RESULTS_DIR, exist_ok=True)

#     images_tensor, labels     = load_cifar10()
#     model, preprocess, device = load_clip_model()
#     embeddings                = generate_embeddings(images_tensor, model, preprocess, device)
#     index                     = build_faiss_index(embeddings)

#     # ── Excel workbook ───────────────────────────────────────────────────────
#     wb = Workbook()

#     # ── Summary sheet ────────────────────────────────────────────────────────
#     summary_ws = wb.active
#     summary_ws.title = "Summary"

#     sum_headers = [
#         "Query #", "Image Filename",
#         "Top-1 Class", "Top-1 Score",
#         "Top-3 Classes (comma-separated)", "Top-3 Scores",
#         "Top-5 Classes (comma-separated)", "Top-5 Scores",
#         f"Top-{TOP_K} Classes (comma-separated)", f"Top-{TOP_K} Scores"
#     ]
#     summary_ws.append(sum_headers)
#     style_header(summary_ws, 1, len(sum_headers))

#     col_widths_summary = [9, 30, 16, 12, 40, 35, 50, 45, 65, 60]
#     for i, w in enumerate(col_widths_summary, 1):
#         summary_ws.column_dimensions[get_column_letter(i)].width = w
#     summary_ws.row_dimensions[1].height = 22

#     # ── Process each query ───────────────────────────────────────────────────
#     print(f"\n[5/5] Running {len(query_paths)} queries...\n")

#     all_rows = []   # collected for per-query sheets

#     for q_idx, img_path in enumerate(query_paths, 1):
#         fname = os.path.basename(img_path)
#         print(f"  [{q_idx:>3}/{len(query_paths)}] {fname}", end=" ... ")

#         try:
#             distances, indices = query_image(img_path, model, preprocess, device, index)
#         except Exception as e:
#             print(f"FAILED ({e})")
#             continue

#         result_labels = [labels[i] for i in indices]
#         result_classes = [CLASSES[l] for l in result_labels]

#         # ── Per-query detail sheet ────────────────────────────────────────────
#         safe_name = os.path.splitext(fname)[0][:28]   # sheet name limit
#         sheet_name = f"Q{q_idx:03d}_{safe_name}"[:31]
#         ws = wb.create_sheet(title=sheet_name)

#         detail_headers = ["Rank", "CIFAR-10 Index", "Class", "Similarity Score"]
#         ws.append(detail_headers)
#         style_header(ws, 1, len(detail_headers))
#         for col_i, width in enumerate([8, 16, 18, 18], 1):
#             ws.column_dimensions[get_column_letter(col_i)].width = width
#         ws.row_dimensions[1].height = 20

#         # title above header
#         ws.insert_rows(1)
#         ws.merge_cells("A1:D1")
#         title_cell = ws["A1"]
#         title_cell.value     = f"Query #{q_idx} — {fname}"
#         title_cell.font      = Font(name="Arial", bold=True, size=12, color="1F4E79")
#         title_cell.alignment = CENTER
#         ws.row_dimensions[1].height = 24

#         for rank, (cifar_idx, cls, dist) in enumerate(
#             zip(indices, result_classes, distances), 1
#         ):
#             ws.append([rank, int(cifar_idx), cls, round(float(dist), 6)])
#             style_body_row(ws, rank + 2, 4, alternate=(rank % 2 == 0))

#         # ── Summary row ───────────────────────────────────────────────────────
#         top1_cls   = result_classes[0]
#         top1_score = round(float(distances[0]), 4)
#         top3_cls   = ", ".join(result_classes[:3])
#         top3_scr   = ", ".join(f"{d:.4f}" for d in distances[:3])
#         top5_cls   = ", ".join(result_classes[:5])
#         top5_scr   = ", ".join(f"{d:.4f}" for d in distances[:5])
#         topk_cls   = ", ".join(result_classes)
#         topk_scr   = ", ".join(f"{d:.4f}" for d in distances)

#         summary_row = [
#             q_idx, fname,
#             top1_cls, top1_score,
#             top3_cls, top3_scr,
#             top5_cls, top5_scr,
#             topk_cls, topk_scr
#         ]
#         summary_ws.append(summary_row)
#         data_row = q_idx + 1
#         style_body_row(summary_ws, data_row, len(sum_headers), alternate=(q_idx % 2 == 0))

#         print(f"Top-1 → {top1_cls} ({top1_score:.4f})")
#         all_rows.append(summary_row)

#     # ── Freeze panes & auto-filter on Summary ────────────────────────────────
#     summary_ws.freeze_panes = "A2"
#     summary_ws.auto_filter.ref = f"A1:{get_column_letter(len(sum_headers))}1"

#     # ── Save Excel ────────────────────────────────────────────────────────────
#     excel_path = os.path.join(RESULTS_DIR, "visual_rag_results.xlsx")
#     wb.save(excel_path)
#     print(f"\n  Excel saved → '{excel_path}'")
#     print(f"  Sheets: 'Summary' + {len(all_rows)} per-query detail sheets")
#     print("\n Done!\n")


# if __name__ == "__main__":
#     main()

# # import os
# # import sys
# # import numpy as np
# # import torch
# # import open_clip
# # import faiss
# # import matplotlib.pyplot as plt
# # import matplotlib.patches as mpatches

# # from torchvision import datasets, transforms
# # from PIL import Image

# # NUM_IMAGES   = 1500
# # BATCH_SIZE   = 64
# # TOP_K        = 10
# # CLIP_MODEL   = "ViT-B-32"
# # PRETRAINED   = "openai"
# # DATA_DIR     = "./cifar_data"

# # CLASSES = [
# #     "airplane", "automobile", "bird", "cat", "deer",
# #     "dog", "frog", "horse", "ship", "truck"
# # ]


# # def load_cifar10(num_images=NUM_IMAGES):
# #     print(f"\n[1/5] Loading CIFAR-10 ({num_images} images)...")
# #     dataset = datasets.CIFAR10(
# #         root=DATA_DIR, train=True, download=True,
# #         transform=transforms.ToTensor()
# #     )
# #     images_tensor, labels = [], []
# #     for idx in range(num_images):
# #         img, label = dataset[idx]
# #         images_tensor.append(img)
# #         labels.append(label)
# #     print(f"    Loaded {len(images_tensor)} images.")
# #     return images_tensor, labels


# # def load_clip_model():
# #     print(f"\n[2/5] Loading CLIP model...")
# #     device = "cuda" if torch.cuda.is_available() else "cpu"
# #     print(f"    Using device: {device}")
# #     model, _, preprocess = open_clip.create_model_and_transforms(
# #         CLIP_MODEL, pretrained=PRETRAINED
# #     )
# #     model = model.to(device).eval()
# #     return model, preprocess, device


# # def tensor_to_pil(img_tensor):
# #     np_img = (img_tensor.permute(1, 2, 0).numpy() * 255).astype(np.uint8)
# #     return Image.fromarray(np_img)


# # def generate_embeddings(images_tensor, model, preprocess, device):
# #     print(f"\n[3/5] Generating embeddings...")
# #     all_embeddings = []
# #     with torch.no_grad():
# #         for start in range(0, len(images_tensor), BATCH_SIZE):
# #             batch = images_tensor[start:start + BATCH_SIZE]
# #             clip_input = torch.stack(
# #                 [preprocess(tensor_to_pil(t)) for t in batch]
# #             ).to(device)
# #             emb = model.encode_image(clip_input)
# #             emb = emb / emb.norm(dim=-1, keepdim=True)
# #             all_embeddings.append(emb.cpu().numpy())
# #             done = min(start + BATCH_SIZE, len(images_tensor))
# #             print(f"    {done}/{len(images_tensor)}", end="\r")
# #     print()
# #     embeddings_np = np.vstack(all_embeddings).astype(np.float32)
# #     print(f"    Shape: {embeddings_np.shape}")
# #     return embeddings_np


# # def build_faiss_index(embeddings):
# #     print(f"\n[4/5] Building FAISS index...")
# #     dim   = embeddings.shape[1]
# #     index = faiss.IndexFlatIP(dim)
# #     index.add(embeddings)
# #     print(f"    Vectors stored: {index.ntotal}")
# #     return index


# # def query_from_image_file(image_path, model, preprocess, device, index, top_k=TOP_K):
# #     """
# #     Load any image from disk, embed it with CLIP, search FAISS.
# #     Supports JPG, PNG, BMP, WEBP — anything PIL can open.
# #     """
# #     print(f"\n[5/5] Querying with image: '{image_path}'...")

# #     if not os.path.exists(image_path):
# #         print(f"    ERROR: File not found — '{image_path}'")
# #         sys.exit(1)

# #     pil_img    = Image.open(image_path).convert("RGB")   
# #     clip_input = preprocess(pil_img).unsqueeze(0).to(device)

# #     with torch.no_grad():
# #         emb = model.encode_image(clip_input)
# #         emb = emb / emb.norm(dim=-1, keepdim=True)
# #         emb = emb.cpu().numpy().astype(np.float32)

# #     distances, indices = index.search(emb, top_k)
# #     return distances[0], indices[0], pil_img


# # def display_results(query_pil, result_tensors, result_labels, distances):
# #     fig, axes = plt.subplots(2, 6, figsize=(18, 6))
# #     fig.suptitle(
# #         f"Visual RAG — Input Image vs Top {TOP_K} Similar CIFAR-10 Images",
# #         fontsize=14, fontweight="bold"
# #     )

# #     axes[0][0].imshow(query_pil)
# #     axes[0][0].set_title("INPUT\n(your image)", fontsize=10, fontweight="bold", color="navy")
# #     axes[0][0].axis("off")
# #     axes[1][0].axis("off")

# #     for rank, (tensor, label, dist) in enumerate(zip(result_tensors, result_labels, distances)):
# #         row = rank // 5
# #         col = (rank %  5) + 1
# #         axes[row][col].imshow(tensor_to_pil(tensor))
# #         axes[row][col].set_title(
# #             f"#{rank+1}  {CLASSES[label]}\nscore: {dist:.3f}",
# #             fontsize=8.5
# #         )
# #         axes[row][col].axis("off")

# #     plt.tight_layout()
# #     plt.savefig("rag_results.png", bbox_inches="tight", dpi=130)
# #     print("    Saved → 'rag_results.png'")
# #     plt.show()


# # def main():
# #     QUERY_IMAGE_PATH = "download1.jpg" #change the image for different querries.

# #     print("=" * 50)
# #     print("  Visual RAG  —  CIFAR-10 + CLIP + FAISS")
# #     print("=" * 50)

# #     images_tensor, labels        = load_cifar10()
# #     model, preprocess, device    = load_clip_model()
# #     embeddings                   = generate_embeddings(images_tensor, model, preprocess, device)
# #     index                        = build_faiss_index(embeddings)
# #     distances, indices, query_pil = query_from_image_file(
# #         QUERY_IMAGE_PATH, model, preprocess, device, index
# #     )

# #     result_tensors = [images_tensor[i] for i in indices]
# #     result_labels  = [labels[i]         for i in indices]

# #     print(f"    Top {TOP_K} results:")
# #     for rank, (idx, dist, lbl) in enumerate(zip(indices, distances, result_labels)):
# #         print(f"      #{rank+1:>2}  idx={idx:>4}  score={dist:.4f}  class={CLASSES[lbl]}")

# #     display_results(query_pil, result_tensors, result_labels, distances)
# #     print("\n Done!\n")


# # if __name__ == "__main__":
# #     main()